import csv
from tempfile import TemporaryFile
import requests as r
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook


class Base:

    def __init__(self) -> None:
        self.file = TemporaryFile(mode="r+")
        self.content = ""
        self.soup = None

    def write_to_temp(self, file) -> bool:

        try:
            self.file.write(file.read().decode())
            self.file.seek(0)
        except:
            return False

        return True

    def _update_dicts(self, main, main_cols, second, second_cols):
        """
        when we use the update method into a dict
        this return None, so we use 'or'
        to use the already updated dict
        """
        result = [
            dict.update(
                {key: None for key in main_cols}
            ) or dict for dict in main.copy() if type(dict).__name__ == "dict"
        ]

        return result + [
            dict.update(
                {key: None for key in second_cols}
            ) or dict for dict in second.copy() if type(dict).__name__ == "dict"
        ]

    def join_dicts(self, dict_one, dict_one_cols, dict_two, dict_two_cols):
        """
        Main function to join dicts from files/url
        will add columns and values from one to other
        data that dosen't exist will be created as NaN
        """

        cols_dif_main = [
            col for col in dict_two_cols if col not in dict_one_cols]
        cols_dif_second = [
            col for col in dict_one_cols if col not in dict_two_cols]

        return self._update_dicts(dict_one, cols_dif_main, dict_two, cols_dif_second)


class IsCsv(Base):

    def parseData(self, index_data):

        data = []
        reader = csv.DictReader(self.file)
        columns = reader.fieldnames
        index = index_data

        for row in reader:
            row_pk = {"pk": index_data}
            row_pk.update(row)
            data.append(row_pk)
            index += 1

        self.file.close()

        return columns, data, index


class isExcel(Base):

    def parseData(self, file, index_data):
        wb = load_workbook(file, read_only=True)

        excel_data = []
        columns = []

        index = index_data

        for wb_sheet in wb.sheetnames:

            sheet = wb[wb_sheet]
            data = []

            header_row = [sheet.cell(row=1, column=col_idx).value or "column-%s" %
                          col_idx for col_idx in range(1, (sheet.max_column or -1) + 1)]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                sheet_dict = {"pk": index}
                sheet_dict.update({header_row[row_index] or "column-%s" % row_index: row[row_index]
                                  for row_index in range(len(header_row))})  # type: ignore
                data.append(sheet_dict)
                index += 1

            if data:
                if excel_data:
                    excel_data = self.join_dicts(
                        excel_data, columns, data, header_row)
                else:
                    excel_data = data

                columns.extend(
                    [col for col in header_row if col not in columns])

        return columns, excel_data, index


class IsUrl(Base):

    def setContent(self, url):
        get = r.get(url)
        self.content = get.text
        self.setParser()

    def setParser(self):
        self.soup = bs(self.content, "html.parser")

    def getData(self, index_data):

        table = self.soup.find_all("table")

        all_table_data = []
        columns = []
        index = index_data

        for i in table:

            table_data = []
            head = i.find("thead")
            body = i.find("tbody")

            if not head:
                continue

            header_row = [i.text.strip() for i in head.find_all("th")]

            body_tr = body.find_all("tr")
            for b_tr in body_tr:
                body_td = b_tr.find_all("td")

                row_pk = {"pk": index}
                row_pk.update({header_row[i]: body_td[i].text.strip()
                              for i in range(len(header_row))})
                table_data.append(row_pk)

                index += 1

            if not table_data:
                continue

            if all_table_data:
                all_table_data = self.join_dicts(
                    all_table_data, columns, table_data, header_row)
            else:
                all_table_data = table_data

            columns.extend(header_row)

        return columns, all_table_data, index
